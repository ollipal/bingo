from typing import List
from pydantic import ValidationError
from schemas import BingoData, Quest, Phrase
from io import BytesIO
from openpyxl import load_workbook

def extract_bingo_data(file_content: bytes) -> tuple[BingoData | None, str | None]:
    """
    Extract and validate data from Excel file content.
    Args:
        file_content: Raw bytes of the Excel file
    Returns:
        A tuple of (validated_data, error) where validated_data is a BingoData instance
        and error is an error message string if any error occurred.
    """
    
    required_sheets = ['quests', 'phrases', 'pattern']
    
    try:
        # Load workbook from bytes
        wb = load_workbook(filename=BytesIO(file_content), read_only=True, data_only=True)
        
        # Check for required sheets
        missing_sheets = [sheet for sheet in required_sheets if sheet not in wb.sheetnames]
        if missing_sheets:
            return None, f"Missing sheets: {', '.join(missing_sheets)}"
        
        # Process each sheet
        processed_data = {
            'quests': [],
            'phrases': [],
            'difficulties': []
        }

        def process_quest_row(headers, row):
            row_data = {}
            types_header = 'types (order matters! first ones are filled first)'
            
            for header, cell in zip(headers, row):
                if header:
                    value = cell.value if cell.value is not None else ''
                    if header == types_header and value:
                        value = [t.strip() for t in str(value).split(';') if t.strip()]
                    row_data[header.replace(types_header, 'types')] = value
            return row_data

        def process_phrase_row(headers, row):
            translations = []
            other_fields = {}
            for header, cell in zip(headers, row):
                value = cell.value if cell.value is not None else ''
                if header.startswith('translation-'):
                    if value and str(value).strip():  # Only include non-empty translations
                        language = header.replace('translation-', '')
                        translations.append({
                            'language': language,
                            'text': str(value).strip()
                        })
                else:
                    other_fields[header] = value
            return {**other_fields, 'translations': translations}

        def process_pattern_sheet(ws):
            """Process pattern sheet as 5x5 grid of difficulty values"""
            difficulties = []
            
            # Process all 5 rows (no header)
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5), start=1):
                # Get first 5 values from the row
                for cell in row[:5]:
                    value = cell.value if cell.value is not None else ''
                    difficulties.append(str(value).strip())

            if len(difficulties) != 25:
                return None, f"Board must contain exactly 25 fields (found {len(difficulties)})"
                
            return difficulties, None

        def is_row_valid(row, headers, required_field=None):
            if not required_field:
                # If no specific field required, check first column
                return row[0].value is not None and str(row[0].value).strip() != ''
            
            # Find the required field in headers and check its value
            for header, cell in zip(headers, row):
                if header == required_field:
                    return cell.value is not None and str(cell.value).strip() != ''
            return False

        def process_sheet(sheet_name, process_row_func, required_field=None):
            ws = wb[sheet_name]
            headers = [str(cell.value) for cell in next(ws.rows) if cell.value is not None]
            
            processed_rows = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if is_row_valid(row, headers, required_field):
                    row_data = process_row_func(headers, row)
                    if row_data:
                        processed_rows.append(row_data)
            return processed_rows

        # Process quests and phrases
        processed_data['quests'] = process_sheet('quests', process_quest_row, 'name')
        processed_data['phrases'] = process_sheet('phrases', process_phrase_row, 'difficulty')

        # Board processing results
        difficulties, error = process_pattern_sheet(wb['pattern'])
        if error:
            return None, error
        processed_data['difficulties'] = difficulties

        # Validate all data at once
        try:
            validated_data = BingoData(
                quests=processed_data['quests'],
                phrases=processed_data['phrases'],
                difficulties=processed_data['difficulties']
            )
            return validated_data, None
        except ValidationError as e:
            # Get the first validation error
            error = e.errors()[0]
            location = error['loc']
            
            if len(location) >= 2:
                data_type = location[0]  # quests, phrases, or difficulties
                index = location[1] if isinstance(location[1], int) else None
                field = '.'.join(str(loc) for loc in location[2:]) if len(location) > 2 else None
                
                error_msg = f"{data_type.capitalize()} "
                if index is not None:
                    error_msg += f"(row {index + 2}) "  # +2 because of header row and 0-based index
                if field:
                    error_msg += f"field '{field}' "
                error_msg += error['msg']
            else:
                error_msg = f"Validation error: {error['msg']}"
            
            return None, error_msg
            
    except Exception as e:
        import traceback
        return None, f"Error: {str(e)}\n{traceback.format_exc()}"